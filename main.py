import openpyxl

interest_sub_list_for_trainers = []


def read_excel_file(keyfile_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(keyfile_path)

    # Assuming the data is in the first sheet (index 0)
    sheet = workbook.worksheets[0]

    columns_to_read = list(range(1, 8)) + list(range(18, 31))

    # Create a list to store the data
    data = []

    # Iterate through the rows and columns to read the data
    for dataRow in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        selected_row = [dataRow[col_idx - 1] for col_idx in columns_to_read]
        data.append(selected_row)

    return data


def get_interest_rate(interest):
    interest_rate = 0
    if interest == "No idea about this":
        interest_rate = 1
    elif interest == "Have Heard about this":
        interest_rate = 2
    elif interest == "Somewhat Interested in this":
        interest_rate = 3
    elif interest == "Very Interested in this":
        interest_rate = 4
    if interest == "I am an expert at this bruh!":
        interest_rate = 5
    return interest_rate


def get_trainers_interest(data):
    interest_list = []
    for row in data:
        person = [row[1]]
        for i in range(10, 20):
            person.append(get_interest_rate(row[i]))
        interest_list.append(person)

    return interest_list


def calculate_matching_values(interest):
    matching_values_list = []

    for trainner in interest_sub_list_for_trainers:
        matching_details = []
        matching_value = 0
        count = 1
        for oneInterest in interest:
            matching_value += abs(oneInterest - trainner[count])
            count += 1
        matching_details.append(trainner[0])
        matching_details.append(matching_value)

        matching_values_list.append(matching_details)

    return matching_values_list


def create_valive_for_matching(participant):
    participant_interest = participant[1:]
    return calculate_matching_values(participant_interest)


def matching_trainers_to_participation(participants):
    matching_list = []
    for participant in participants:
        participant_details = [participant[0], create_valive_for_matching(participant)]
        matching_list.append(participant_details)

    return matching_list


def sort_list_by_second_value(data_list):
    # Sort the list in ascending order based on the second value in each sublist
    sorted_list = sorted(data_list, key=lambda x: x[1])
    return sorted_list


def sort_list_trainners_list(data_list):
    sorted_list = []
    for data in data_list:
        sorted_list.append([data[0], sort_list_by_second_value(data[1])])

    return sorted_list


def create_excal_file(data_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matching"
    ws.append(["email", "matching first", "matching second ", "matching third"])

    for data in data_list:
        print()
        ws.append([str(data[0]), str(data[1][0]), str(data[1][1]), str(data[1][1])])

    wb.save("Matching.xlsx")


if __name__ == "__main__":
    # get the user input using cmd
    # Trainers Form - 3K (Responses).xlsx
    # file_path = input("Enter the path to the excel file: ")
    # start_row = int(input("Enter the start row: "))
    # end_row = int(input("Enter the end row: "))

    file_path_for_trainers = "Trainers Form - 3K (Responses).xlsx"
    file_path_for_participant = "Trainers Form - 3K (Responses).xlsx"
    start_row = 3
    end_row = 19

    excel_data_for_trainers = read_excel_file(file_path_for_trainers)
    excel_data_for_participant = read_excel_file(file_path_for_participant)

    interest_sub_list_for_trainers = get_trainers_interest(excel_data_for_trainers)
    interest_sub_list_for_participant = get_trainers_interest(excel_data_for_participant)

    final_matching_list = matching_trainers_to_participation(interest_sub_list_for_participant)

    final_matching_list_sorted = sort_list_trainners_list(final_matching_list)

    out_put_data_set = []

    for match in final_matching_list_sorted:
        out_put_data_set.append([match[0], match[1][0:3]])

    create_excal_file(out_put_data_set)
